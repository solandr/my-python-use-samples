#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Sep  2 16:03:52 2022

@author: Amdrew Solowey
"""
from openpyxl import load_workbook
import re
import pytz
import requests
from urllib.parse import urlencode
from urllib.parse import urlparse
from urllib.parse import parse_qs
from datetime import datetime
import os
import glob
from os.path import isfile, join
from os import listdir, scandir
import zipfile

class AccPos:
    ''' Acceptance position definition'''
    def __init__(self, position, updated = False):
        self.__position = position
        self.__updated = updated
        
    @property
    def position(self):
        return self.__position
    @position.setter
    def position(self, position):
        self.__position = position
        
    @property
    def updated(self):
        return self.__updated
    @updated.setter
    def updated(self, updated):
        self.__updated = updated

class Fhlp:
    
    @staticmethod
    def fnum(val):
        try:
            if (val == None):
                return 0
            return float(val)
        except ValueError:
            if isinstance(val, str):
                valx = val.replace(',', '.').replace("р", "").replace(" ", "")
                try:
                    return float(valx)
                except ValueError:
                    print(f"Unknown number {val}")
                    return 0
                else:
                    print(f"Unknown number {val}")
                    return 0
    @staticmethod
    def fint(val):
        try:
            if (val == None):
                return 0
            return int(val)
        except ValueError:
            if isinstance(val, str):
                valx = val.replace(',', '.').replace(" ", "")
                try:
                    return int(valx)
                except ValueError:
                    print(f"Unknown number {val}")
                    return 0
                else:
                    print(f"Unknown number {val}")
                    return 0

    @staticmethod
    def ind(lst, item):
        try:
            return lst.index(item)
        except ValueError:
            return -1
            
    @staticmethod
    def kint(val):
             if (val == None):
                 return 0
             if isinstance(val, str):
                 valx = val.replace(',', '.').replace(" ", "").replace("-", "")
                 try:
                     return int(valx)
                 except ValueError:
                     print(f"kint1: Unknown number {val}")
                     return 0
                 else:
                     print(f"kint2: Unknown number {val}")
                     return 0

    @staticmethod
    def alnum(val):
        if isinstance(val, str):
            return ''.join(filter(str.isalnum, val))
        return ''.join(filter(str.isalnum, f"{val}"))
    
    @staticmethod
    def split(val):
        if val == None:
            return ""
        if isinstance(val, str):
            return val.strip()
        return f'{val}'.strip()
    
    @staticmethod
    def strip(val):
        if val == None:
            return ""
        if isinstance(val, str):
            return val.strip()
        return f'{val}'.strip()

    @staticmethod
    def create_dir_if_not_exist(path):
        try:
            os.makedirs(path, exist_ok = True)
        except OSError as err:
            print(f"Current dir: {os.getcwd()}")
            print(f"Directory '{path}' can not be created.\n{err}")
            exit()

    @staticmethod
    def is_not_empty_folder(path):
        if isfile(path):
            return False
        for f in scandir(path):
            if f.is_file():
                return True
        return False
    
    @staticmethod
    def get_subdir_list(source_path):
        files_info = []
        try:
            files_info = [f for f in listdir(source_path) if Fhlp.is_not_empty_folder(join(source_path, f))]
        except FileNotFoundError:
            print(f"The directory {source_path} does not exist")
        except PermissionError:
            print(f"Permission denied to access the directory {source_path}")
        except OSError as e:
            print(f"An OS error occurred: {e}")
        return files_info
        
    @staticmethod
    def zipdir(path, ziph):
        for p in scandir(path):
            if p.is_file():
                ziph.write(p.path, 
                           os.path.relpath(join(path, p.name), 
                                           join(path, '..')))
    
    @staticmethod
    def make_archives(source_path, target_path):
        dir_list = Fhlp.get_subdir_list(source_path)
        if len(dir_list) > 0:
            Fhlp.create_dir_if_not_exist(target_path)
            for d in dir_list:
                d_path = join(target_path, f"{d}.zip")
                with zipfile.ZipFile(d_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    Fhlp.zipdir(join(source_path, d), zipf)
    
    @staticmethod
    def get_last_file(pathF, mask, mask_tmpl, ext, tt):
        if tt == 1:
        # НАЦЕНКИ - ИНФ Excel 12,04,2023.xlsx
            ftmpl = fr"{mask_tmpl}.*\s(?P<d>\d+)[\,\.](?P<m>\d+)[\,\.](?P<y>\d+)\.{ext}"
        else:
        # F99_price_(20230617_18-46).xlsx
            ftmpl = fr"{mask_tmpl}.*(?P<y>\d{{4}})(?P<m>\d{{2}})(?P<d>\d{{2}})\_(?P<hh>\d{{2}})-(?P<mm>\d{{2}}).?\.{ext}"
            # ftmpl = fr"{mask}.*(?P<y>\d{{4}})(?P<m>\d{{2}})(?P<d>{{2}})(\_(?P<hh>{{2}})-(?P<mm>\d{{2}}))?.?\.{ext}"
        pathX = join(pathF, mask)
        xlsxFiles = glob.glob(fr'{pathX}*.{ext}')
        
        p = re.compile(ftmpl)
        
        utc=pytz.UTC
        last_d = utc.localize(datetime.min)
        res = None
        for f in xlsxFiles:
            # test_str = u"F99/F99_price_(20230605_18-40).xlsx"
            match = re.search(p, f)

            if match != None:
                ln = len(match.groupdict())
                if match != None and ln > 2:
                    d = int(match.group('d'))
                    m = int(match.group('m'))
                    y = int(match.group('y'))
                    if ln > 3:
                        hh = int(match.group('hh'))
                        mm = int(match.group('mm'))
                        date = utc.localize(datetime(year = y, month = m, day = d, hour = hh, minute = mm))
                    else:
                        date = utc.localize(datetime(year = y, month = m, day = d))
                    if date > last_d:
                        res = f
                        last_d = date
        return (res, last_d)
 
    @staticmethod
    def get_row_key(brand, articule):
        return f"{Fhlp.alnum(brand)}|{Fhlp.alnum(articule)}"
 
    @staticmethod
    def get_xlsx_content(pathF, 
                         used_pairs = {"код":"number", "бренд": "brand", "количество": "quantity",
                          "наша цена": "supBuyPrice", "штрих-код": "barcodes", "наименование": "descr"},
                         key_field = "дата поставки", sub_key = "код", sub_key_bc = "бренд",
                         first_row = 1, brand_subst = None):
        
        def get_key_by_val(ditems, vx):
            for k, v in ditems:
                if v == vx:
                    return k
            return None
        
        def get_dict_row(r, n, col_names):
            cur_row = {}
            i = 0
            for cell in r:
                cur_row[col_names[i]] = cell.value
                i = i + 1
                if i == n:
                    break
            return cur_row
        
        price_gname = "supBuyPrice"
        quantity_gname = "quantity"
        ditems = used_pairs.items()
        pk = get_key_by_val(ditems, price_gname)
        qk = get_key_by_val(ditems, quantity_gname)
        if (pk == None) or (qk == None):
            raise Exception(f"Invalid data structure.\nsupBuyPrice <-> {pk}\nquantity <-> {pk}\n")
            
        wb = load_workbook(pathF, read_only=True)
        ws = wb[wb.sheetnames[0]]
        col_names = Fhlp.get_sheet_col_names(ws)

        inp_row_count = ws.max_row
        rd = {}
        n = len(col_names)
        # swTest = False
        for r in ws.iter_rows(min_row=first_row + 1, max_row=inp_row_count):
            cur_row = get_dict_row(r, n, col_names)
            
            articule = cur_row[sub_key]

            price = Fhlp.fnum(cur_row[pk])
            qt = Fhlp.fnum(cur_row[qk])
            
            if (price <= 0) or (qt <= 0):
                continue


            # shipment_date = get_date(row[key_field])
            # shipment_dict = rd.get(shipment_date, None)
            # if (shipment_dict == None):
            #     shipment_dict = {}
            #     rd[shipment_date] = shipment_dict
                
            # articule = cur_row[sub_key]
            
            brand = cur_row[sub_key_bc]

            key = Fhlp.get_row_key(brand, articule)
            
            # if articule == "8450007313":
            #     print(f"get_xlsx_content: key = '{key}'; articule = '{articule}'; brand = '{articule}'; price = {price}.\n" )
            #     swTest = True
            # else:
            #     swTest = False
                
            good_row = {}
            
            for (k, v) in ditems:
                good_row[v] = cur_row[k]
                    
            if brand_subst != None:
                good_row["brand"] = brand_subst

            good_row[price_gname] = price
            good_row[quantity_gname] = qt
            bc = good_row["barcodes"]
            barcodes = set(bc.split(" ")) if bc != None else set()
            good_row["barcodes"] = barcodes

            if key in rd:
                sp_gr = rd[key]
                if price in sp_gr:
                    pg = sp_gr[price].position
                    pg["barcodes"].update(barcodes)
                    pg["quantity"] += qt
                    # if swTest:
                    #     print(f"get_xlsx_content: set quantity = {pg['quantity']}\n")
                else:
                    sp_gr[price] = AccPos(good_row) # check updated False if not
                    # if swTest:
                    #     print(f"get_xlsx_content: add dict for price = {price}\n")
            else:
                rd[key] = {price: AccPos(good_row)} # check updated False if not
                # if swTest:
                #     print(f"get_xlsx_content: add key = '{key}'; price = {price}.\n")

        return rd


    @staticmethod
    def get_sheet_col_names(sheet, first_row = 1):
        col_names = []
        for row in sheet.iter_rows(min_row=first_row, max_row=first_row):
            for cell in row:
                v = cell.value
                if v == None:
                    break
                col_names.append(v)
        return col_names


class Fdi:
    
    comp_dict = {}
    tp_dict = {}
    mrj_dict = {}
    v_list = []
    mrj_dict_1 = {}
    v_list_1 = []
    
    dict_v_lists = {}
    
    base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources'
    public_key = 'https://disk.yandex.by/d/KJBj-GnrR_MD6w'
    ftmpl = r"НАЦЕНКИ.*\s(?P<d>\d+)[\,\.](?P<m>\d+)[\,\.](?P<y>\d+)\.xlsx"
    
    @classmethod
    def get_cur_marj_file(cls):
        xlsxFiles = glob.glob('НАЦЕНКИ*.xlsx')
        valid = re.compile(cls.ftmpl)
        utc=pytz.UTC
        last_d = utc.localize(datetime.min)
        res = None
        for f in xlsxFiles:
            match = valid.match(f)
            if match != None and len(match.groupdict()) > 2:
                d = match.group('d')
                m = match.group('m')
                y = match.group('y')
                date = utc.localize(datetime.fromisoformat(f'{y}-{m}-{d}'))
                if date > last_d:
                    res = f
                    last_d = date
        return (res, last_d)
    
    @classmethod
    def get_marj_path(cls):
        cfile, cdate = cls.get_cur_marj_file()
        
        valid = re.compile(cls.ftmpl)
        utc=pytz.UTC
        # Получаем загрузочную ссылку
        ypath = '/';
        
        yfields = '_embedded.items.name,_embedded.items.type';
         
        ylimit = 100;
        
        p1 = urlencode(dict(path=ypath))
        p2 = urlencode(dict(public_key=cls.public_key))
        final_url = f"{cls.base_url}??{p1}&{p2}&fields={yfields}&limit={ylimit}"
        
        response = requests.get(final_url)
        js = response.json()
        items = js["_embedded"]["items"]
        pk = js["_embedded"]["public_key"]
        p0 = urlencode(dict(public_key=pk))
        last_d = cdate
        res = None
        for it in items:
            if it['type'] != 'file' : continue
            it_name = it["name"]
            match = valid.match(it_name)
            if match != None and len(match.groupdict()) > 2:
                d = match.group('d')
                m = match.group('m')
                y = match.group('y')
                date = utc.localize(datetime.fromisoformat(f'{y}-{m}-{d}'))
                if date > last_d:
                    res = it_name
                    last_d = date
 
    # exists_for_real = os.path.isfile(file_path) and os.path.getsize(file_path) > 0
        
        if res == None:
            return (cfile, cdate)
        
        p02 = urlencode(dict(path=f'/{res}'))
        
        file_url = f"{cls.base_url}/download?{p02}&{p0}"
        response = requests.get(file_url)
        
        js = response.json()
        
        download_url = js['href']
        
        parsed_url = urlparse(download_url)
        file_name = parse_qs(parsed_url.query)['filename'][0]
        print(f"get НАЦЕНКИ from file: {file_name}")
        
        file_nameR = file_name
        exists_for_real = os.path.isfile(file_nameR) and os.path.getsize(file_nameR) > 0
        
        if exists_for_real:
            return file_nameR
        
        download_response = requests.get(download_url)
        open(file_nameR, 'wb').write(download_response.content)
                    
        return file_name, last_d
    
    @classmethod
    def load_marjs(cls,fn):
        wb = load_workbook(fn)
        df = {}
        for ws in wb.worksheets:
            inp_row_count = ws.max_row
            rows = []
            for r in ws.iter_rows(min_row=3, max_row=inp_row_count):
                d = r[2].value
                if d == None or d == "":
                    continue
                v = r[3].value
                if v == None or v == "":
                    continue
                if str(d).startswith("9999"):
                    d = -1
                rows.append((d , v))
            
            df[ws.title] = rows

        return df

# НАЦЕНКИ - ИНФ Excel 30,10,2022.xlsx
    @classmethod
    def init_data(cls):
        fn, fd = cls.get_marj_path()
        if fn != None:
            cls.dict_v_lists = cls.load_marjs(fn)
        
        wb = load_workbook("helper.xlsx")
        ws = wb["fbr"]
        
    #    for cell in ws['A']:
        for row in range(2, ws.max_row + 1):
          key = ws.cell(row, 1).value
          value = ws.cell(row, 2).value
          cls.comp_dict[key] = value
        
        ws = wb["cats"]
        for row in range(2, ws.max_row + 1):
          key = ws.cell(row, 1).value
          value = ws.cell(row, 2).value
          cls.tp_dict[key] = value
          
        for row in range(2, ws.max_row + 1):
          key = ws.cell(row, 5).value
          value = ws.cell(row, 6).value
          if (key is None) : break
          cls.mrj_dict[key] = Fhlp.fnum(value)
          
        for row in range(2, ws.max_row + 1):
          key = Fhlp.fnum(ws.cell(row, 11).value)
          value = Fhlp.fnum(ws.cell(row, 12).value)
          if ((key is None) or (key == 0)) : break
          cls.v_list.append((key,value))
      
        for row in range(2, ws.max_row + 1):
          key = ws.cell(row, 17).value
          value = ws.cell(row, 18).value
          if (key is None) : break
          cls.mrj_dict_1[key] = Fhlp.fnum(value)
          
        for row in range(2, ws.max_row + 1):
          key = Fhlp.fnum(ws.cell(row, 14).value)
          value = Fhlp.fnum(ws.cell(row, 15).value)
          if ((key is None) or (key == 0)) : break
          cls.v_list_1.append((key,value))

    @classmethod
    def get(cls, f_type):
        if len(cls.comp_dict) == 0:
            cls.init_data()
            
        if f_type in cls.dict_v_lists:
            if f_type == "F99-5":
                return (cls.comp_dict, cls.tp_dict, cls.mrj_dict_1, cls.dict_v_lists[f_type])
            else:
                return (cls.comp_dict, cls.tp_dict, cls.mrj_dict, cls.dict_v_lists[f_type])
            
        if f_type == "F99-5":
            return (cls.comp_dict, cls.tp_dict, cls.mrj_dict_1, cls.v_list_1)
        return (cls.comp_dict, cls.tp_dict, cls.mrj_dict, cls.v_list)
