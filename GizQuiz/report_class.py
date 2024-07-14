from datetime import datetime
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import re
from copy import copy
import pandas as pd
import numpy as np

class report:
    def __init__(self, template_name, template_path, report_dir) -> None:
        self.template_path = template_path
        self.template_name = template_name
        self.report_dir = report_dir
        self.decor_values = {
            'Form22': self.decor_form22,
            'region': self.decor_region
            }
        self.field_values = {'period': 'весь период', 'what':'Брестская область', 'date':datetime.now().strftime('%d-%m-%Y')}

    @classmethod
    def decor_region(cls, region_id):
        if region_id == 1: return 'Брестская обл.'
        if region_id == 2: return 'Витебская обл.'
        if region_id == 3: return 'Гомельская обл.'
        if region_id == 4: return 'Гродненская обл.'
        if region_id == 6: return 'Минская обл.'
        if region_id == 7: return 'Могилевская обл.'
        return ''

    @classmethod
    def decor_form22(cls, f22):
        if f22 == 1:
          return '01 - сельскохозяйственные организации, использующие предоставленные им земли для ведения сельского хозяйства, в том числе в исследовательских и учебных целях, а также для ведения подсобного хозяйства'
        if f22 == 2:
          return '02 - сельскохозяйственные организации Министерства сельского хозяйства и продовольствия Республики Беларусь'
        if f22 == 3:
          return '03 - крестьянские (фермерские) хозяйства'
        if f22 == 5:
          return '05 - граждане, использующие земельные участки для строительства и (или) обслуживания жилого дома'
        if f22 == 6:
          return '06 - граждане, использующие земельные участки для ведения личного подсобного хозяйства'
        if f22 == 7:
          return '07 - граждане, использующие земельные участки для садоводства и дачного строительства'
        if f22 == 8:
          return '08 - граждане, использующие земельные участки для огородничества'
        if f22 == 9:
          return '09 - граждане, использующие земельные участки для сенокошения и выпаса сельскохозяйственных животных'
        if f22 == 10:
          return '10 - граждане, использующие земельные участки для иных сельскохозяйственных целей'
        if f22 == 11:
          return '11 - граждане, использующие земельные участки для иных несельскохозяйственных целей'
        if f22 == 12:
          return '12 - промышленные организации'
        if f22 == 13:
          return '13 - организации железнодорожного транспорта'
        if f22 == 14:
          return '14 - организации автомобильного транспорта'
        if f22 == 15:
          return '15 - организации Вооруженных Сил Республики Беларусь, воинских частей, военных учебных заведений и других войск и воинских формирований Республики Беларусь'
        if f22 == 16:
          return '16 - организации воинских частей, военных учебных заведений и других войск и воинских формирований иностранных государств'
        if f22 == 17:
          return '17 - организации связи, энергетики, строительства, торговли, образования, здравоохранения и иные землепользователи'
        if f22 == 18:
          return '18 - организации природоохранного, оздоровительного, рекреационного и историко - культурного назначения'
        if f22 == 19:
          return '19 - заповедники, национальные парки и дендрологические парки'
        if f22 == 20:
          return '20 - организации, ведущие лесное хозяйство'
        if f22 == 21:
          return '21 - организации, эксплуатирующие и обслуживающие гидротехнические и иные водохозяйственные сооружения'
        if f22 == 22:
          return '22 - земли, земельные участки, не предоставленные землепользователям'
        if f22 == 23:
          return '23 - земли общего пользования в населенных пунктах, садоводческих товариществах и дачных кооперативах, а также земельные участки, используемые гражданами'
        if f22 == 24:
          return '24 - иные земли общего пользования за пределами границ населенных пунктов'
        return f"{f22}: ..."  

    def decor_period(self, dates):
        if isinstance(dates, (list, tuple)):
            d1, d2 = dates
            if d1 == None:
                if d2 == None:
                    period = 'за весь период'
                else:
                    period = f'до {d2.strftime("%d %b %Y")}'
            else:
                if d2 == None:
                    period = f'с {d1.strftime("%d %b %Y")}'
                else:
                    period = f'с {d1.strftime("%d %b %Y")} до {d2.strftime("%d %b %Y")}'
        else: 
            return 'за весь период'

    def get_fields(self, worksheet, number_of_rows, number_of_columns):
        tmpl = r"\{([^\{\}]+)\}"
        valid = re.compile(tmpl)
        for i in range(number_of_columns + 1):
            for k in range(number_of_rows + 1):
                cell = worksheet[get_column_letter(i+1)+str(k+1)]
                cellValue = str(cell.value)
                lst = valid.findall(cellValue)
                if len(lst) > 0:
                    for field in lst:
                        fv = self.get_field_value(field)
                        new_value = cellValue.replace("{" + field + "}", fv)
                        cell.value = new_value

    def set_field_values(self, values):
        self.field_values = values
        if values == None or len(values) == 0:
           return
        what = self.get_field_value('what')
        region_id = self.get_field_value('region_id')
        if what == '':
            values['what'] = self.decor_region(region_id)
        region = self.get_field_value('region')
        if region == '':
            values['region'] = self.decor_region(region_id)
        period = self.get_field_value('period')
        if period == '':
            dates = self.get_field_value('dates')
            values['period'] = self.decor_period(dates)

    def get_field_value(self, field):
        return self.field_values.get(field, "") if self.field_values != None else ""
    
    def get_row_fileds(self, worksheet, number_of_rows):
        pref = '>d>'
        dp = {}
        sw = False
        row_ind = 1
        col_ind = 1
        for row in worksheet.iter_rows(min_row=1, max_row= number_of_rows):
            for cell in row:
                if cell.value == None:
                    continue
                vs = str(cell.value)
                if vs.startswith(pref):
                    col_val = vs.removeprefix(pref)
                    dp[col_val] = cell.column, cell.col_idx
                    cell.value = col_val
                    sw = True
                    row_ind = cell.row
                    continue
                if sw and (len(vs) > 0):
                    dp[vs] = cell.column,cell.col_idx
            if sw:
                return dp, row_ind, col_ind
        return None, None, None
    
    def get_total_row_ind(self, worksheet, number_of_rows):
        pref = '>s>'
        dp = {}
        sw = False
        for row in worksheet.iter_rows(min_row=1, max_row= number_of_rows):
            for cell in row:
                if cell.value == None:
                    continue
                vs = str(cell.value)
                if vs.startswith(pref):
                    col_val = vs.removeprefix(pref)
                    cell.value = col_val
                    return cell.row
        return None

    def insert_row(self, ws, row, row_ind, col_ind):
        for cell in row:
           if isinstance(cell, (int, float, complex)) and cell != 0:
              ws.cell(row_ind, col_ind).value  = cell
           col_ind += 1

    def insert_total_rows(self, totals):
        pass
    
    @classmethod
    def copy_row_styles(cls, ws, row_src_ind, row_tar_ind):
       row_src = ws[row_src_ind]
       row_tar = ws[row_tar_ind]
       for cs in row_src:
          if cs.has_style:
             ct = row_tar[cs.col_idx-1]
             ct.font = copy(cs.font)
             ct.border = copy(cs.border)
             ct.fill = copy(cs.fill)
             ct.number_format = copy(cs.number_format)
             ct.protection = copy(cs.protection)
             ct.alignment = copy(cs.alignment)
             
    def insert_header_row(self, ws, row_ind, col_ind, val, add = True):
        if add:
            row_ind1 = row_ind + 1
            ws.insert_rows(row_ind1, 1)
        ws.cell(row_ind1, col_ind).value  = val
        if add:
            self.copy_row_styles(ws, row_ind, row_ind1)
        return row_ind1

    def run(self):
        now = datetime.now()
        curtimeS = now.strftime('%Y%m%d_%H-%M')
        self.report_path = self.prepare_result_file(self.template_path, self.report_dir, self.template_name, curtimeS)
        wb = load_workbook(self.report_path)
        ws = wb.active
        number_of_rows = ws.max_row
        number_of_columns = ws.max_column
        self.get_fields(ws, number_of_rows, number_of_columns)
        row_column_inds, row_ind, col_ind = self.get_row_fileds(ws, number_of_rows)
        if (row_column_inds != None):
            row_ind1 = self.insert_header_row(ws, row_ind + 1, col_ind, str.upper(self.get_field_value('region')))
            
        df = self.get_field_value('cd1')
        df1 = self.get_field_value('cd2')
        
        row_ind = row_ind1 + 1
        for index, row in df.iterrows():
            rayon = str(index[1])
            row_ind = row_ind1
            row_ind1 = self.insert_header_row(ws, row_ind1, col_ind, rayon)
            self.insert_row(ws, row, row_ind1, col_ind + 2)
            dx = df1.query("Rayon == @rayon")
            for index1, row1 in dx.iterrows():
                form22 = self.decor_form22(int(index1[1]))
                row_ind1 = self.insert_header_row(ws, row_ind1, col_ind+1, form22)
                self.insert_row(ws, row1, row_ind1, col_ind + 2)

        number_of_rows = ws.max_row
        ind_total = self.get_total_row_ind(ws, number_of_rows)
        if ind_total > 0:
            df3 = self.get_field_value('cd3')
            row_total = df3.loc[df3.index[0]]
            self.insert_row(ws, row_total, ind_total, col_ind + 2)
            df4 = self.get_field_value('cd4')
            ind_total1 = ind_total + 1
            for index4, row4 in df4.iterrows():
                form22 = self.decor_form22(int(index4[1]))
                ind_total1 = self.insert_header_row(ws, ind_total1, col_ind+1, form22)
                self.insert_row(ws, row4, ind_total1, col_ind + 2)


        wb.save(self.report_path)
        return self.report_path


    @classmethod
    def prepare_result_file(cls, src, dir_name, file_name, timeStr):
        file_path = f"{dir_name}\\{file_name} {timeStr}.xlsx"
        path = os.path.dirname(file_path)
        try:
            os.makedirs(path, exist_ok = True)
        except OSError as err:
            print(f"Current dir: {os.getcwd()}")
            print(f"Directory '{path}' can not be created.\n{err}")
            exit()
        shutil.copy(src, file_path)
        return file_path

